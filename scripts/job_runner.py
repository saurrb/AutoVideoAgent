from __future__ import annotations

import argparse
import importlib.util
import json
import shutil
import sqlite3
import subprocess
import sys
import uuid
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from autovideo.services.config_loader import load_page_config  # noqa: E402
from autovideo.services.state_store import connect  # noqa: E402
from autovideo.services.structured_log import JsonLogger  # noqa: E402
from autovideo.services.telegram_notify import load_dotenv, reel_status_message, send_telegram  # noqa: E402

_daily_path = PROJECT_ROOT / "scripts" / "daily_ui_batch_schedule.py"
_spec = importlib.util.spec_from_file_location("daily_ui_batch_schedule", _daily_path)
if _spec is None or _spec.loader is None:
    raise RuntimeError(f"Unable to load scheduler module: {_daily_path}")
_daily = importlib.util.module_from_spec(_spec)
sys.modules["daily_ui_batch_schedule"] = _daily
_spec.loader.exec_module(_daily)
PagePlan = _daily.PagePlan
_load_plans = _daily._load_plans
_parse_manifest_from_stdout = _daily._parse_manifest_from_stdout
_schedule_dt_for_slot = _daily._schedule_dt_for_slot
_upload_and_schedule_with_retry = _daily._upload_and_schedule_with_retry


@dataclass
class PendingItem:
    page_key: str
    slot: str
    run_date: str
    target_dt: datetime
    asset_id: str
    generator_type: str
    predicted_content_id: int
    reservation_id: str


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _cleanup_stale_processes(max_age_min: int = 30) -> int:
    ps = rf"""
$cutoff = (Get-Date).AddMinutes(-{max_age_min})
$procs = Get-CimInstance Win32_Process | Where-Object {{
  ($_.Name -in @('python.exe','ffmpeg.exe','grok.exe')) -and
  ($_.CreationDate -lt $cutoff) -and
  (
    ($_.CommandLine -like '*AutoVideoAgent*') -or
    ($_.CommandLine -like '*daily_ui_batch_schedule.py*') -or
    ($_.CommandLine -like '*create_and_post_reel.py*') -or
    ($_.CommandLine -like '*autovideo.app.cli*')
  )
}}
$count = 0
foreach($p in $procs) {{
  try {{ taskkill /PID $p.ProcessId /F | Out-Null; $count += 1 }} catch {{}}
}}
Write-Output $count
"""
    proc = subprocess.run(["powershell", "-NoProfile", "-Command", ps], capture_output=True, text=True)
    if proc.returncode != 0:
        return 0
    try:
        return int((proc.stdout or "0").strip().splitlines()[-1])
    except Exception:
        return 0


def _cleanup_old_runs(days_to_keep: int = 7) -> int:
    runs = PROJECT_ROOT / "runs"
    if not runs.exists():
        return 0
    cutoff = datetime.now() - timedelta(days=days_to_keep)
    removed = 0
    for p in runs.iterdir():
        try:
            mtime = datetime.fromtimestamp(p.stat().st_mtime)
            if mtime < cutoff and p.is_dir():
                shutil.rmtree(p, ignore_errors=True)
                removed += 1
        except Exception:
            continue
    return removed


def _read_json(path: Path, default: Any) -> Any:
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        pass
    return default


def _build_idempotency_key(page_key: str, run_date: str, slot: str, content_id: int) -> str:
    return f"{page_key}:{run_date}:{slot}:{content_id}"


def _probe_video(video_path: Path) -> tuple[float, bool]:
    cmd = ["ffprobe", "-v", "error", "-show_entries", "format=duration", "-show_streams", "-of", "json", str(video_path)]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        return 0.0, False
    data = json.loads(proc.stdout or "{}")
    duration = float((data.get("format") or {}).get("duration") or 0.0)
    streams = data.get("streams") or []
    has_audio = any(str(s.get("codec_type", "")) == "audio" for s in streams)
    return duration, has_audio


def _kill_process_tree(pid: int) -> None:
    try:
        subprocess.run(["taskkill", "/PID", str(pid), "/T", "/F"], capture_output=True, text=True, timeout=20)
    except Exception:
        pass


def _run_render(
    page_key: str,
    heartbeat: Any | None = None,
    timeout_seconds: int = 900,
    slot: str = "",
    target_dt: datetime | None = None,
) -> tuple[Path, dict[str, Any], str]:
    if page_key == "dragon_cinema":
        cmd = [sys.executable, str(PROJECT_ROOT / "scripts" / "generate_dragon_chain_reel.py"), "--page", page_key]
        if slot:
            cmd.extend(["--slot", slot])
        if target_dt is not None:
            cmd.extend(["--schedule-date", target_dt.strftime("%Y-%m-%d")])
    elif page_key == "page4_relationship":
        cmd = [sys.executable, str(PROJECT_ROOT / "scripts" / "generate_page4_reel.py")]
    else:
        cmd = [sys.executable, str(PROJECT_ROOT / "scripts" / "create_and_post_reel.py"), "--page", page_key, "--dry-run"]
    proc = subprocess.Popen(
        cmd,
        cwd=PROJECT_ROOT,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    started = time.time()
    next_beat = 10.0
    while proc.poll() is None:
        elapsed = time.time() - started
        if elapsed >= float(timeout_seconds):
            _kill_process_tree(proc.pid)
            raise RuntimeError(f"Render watchdog timeout after {int(elapsed)}s for page={page_key}")
        if heartbeat and elapsed >= next_beat:
            try:
                heartbeat(round(elapsed, 1))
            except Exception:
                pass
            next_beat += 10.0
        time.sleep(0.5)
    stdout, stderr = proc.communicate()
    if proc.returncode != 0:
        raise RuntimeError(f"Render failed:\n{stdout}\n{stderr}")
    manifest = _parse_manifest_from_stdout(stdout)
    payload = json.loads(manifest.read_text(encoding="utf-8"))
    if page_key == "dragon_cinema":
        caption = str(payload.get("caption", "")).strip()
        hashtags = str(payload.get("hashtags", "")).strip()
        caption = f"{caption}\n\n{hashtags}".strip() if (caption or hashtags) else ""
        if "output_mp4" not in payload and "final_mp4" in payload:
            payload["output_mp4"] = payload["final_mp4"]
    elif page_key == "page4_relationship":
        caption = str(payload.get("caption", "")).strip()
    else:
        caption_file = manifest.with_suffix(".caption.txt")
        caption = caption_file.read_text(encoding="utf-8").strip() if caption_file.exists() else ""
    return manifest, payload, caption


def _grok_dependent(page_key: str) -> bool:
    if page_key in {"dragon_cinema", "page4_relationship"}:
        return True
    try:
        cfg = load_page_config(PROJECT_ROOT, page_key).profile
        return bool(cfg.get("render", {}).get("use_scene_prompt_grok", False))
    except Exception:
        return False


def _classify_generation_error(err: str) -> str:
    s = (err or "").lower()
    if (
        "rate limit" in s
        or "exceeded" in s
        or "429" in s
        or "quota" in s
        or "spending-limit" in s
        or "run out of credits" in s
        or "need a grok subscription" in s
        or "403 forbidden" in s
    ):
        return "grok_rate_limit"
    if "timed out" in s or "timeout" in s or "connection" in s:
        return "transient"
    return "hard_failed"


def _state_set(conn: sqlite3.Connection, job_id: str, state: str, **kwargs: Any) -> None:
    now = _now_iso()
    fields = ["state = ?", "updated_at = ?"]
    vals: list[Any] = [state, now]
    for k, v in kwargs.items():
        fields.append(f"{k} = ?")
        vals.append(v)
    vals.append(job_id)
    conn.execute(f"UPDATE reel_jobs SET {', '.join(fields)} WHERE job_id = ?", vals)
    conn.commit()


def _insert_job(conn: sqlite3.Connection, *, job_id: str, page_key: str, run_date: str, slot: str) -> None:
    now = _now_iso()
    # Use a unique temporary idempotency key at queue/render start.
    # Real deterministic key is set later after content_id is known.
    temp_idem = f"pending:{job_id}"
    conn.execute(
        "INSERT INTO reel_jobs(job_id,page_key,run_date,slot,content_id,idempotency_key,state,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
        (job_id, page_key, run_date, slot, 0, temp_idem, "queued", now, now),
    )
    conn.commit()


def _job_exists_by_idem(conn: sqlite3.Connection, idem: str) -> bool:
    row = conn.execute("SELECT 1 FROM reel_jobs WHERE idempotency_key=?", (idem,)).fetchone()
    return bool(row)


def _db_item_used(conn: sqlite3.Connection, page_key: str, item_id: int) -> bool:
    row = conn.execute("SELECT used FROM content_items WHERE page_key=? AND item_id=?", (page_key, item_id)).fetchone()
    if bool(row and int(row[0]) == 1):
        return True
    row2 = conn.execute("SELECT used FROM used_state WHERE page_key=? AND content_id=?", (page_key, item_id)).fetchone()
    return bool(row2 and int(row2[0]) == 1)


def _predict_next_content_ids(conn: sqlite3.Connection, page_key: str, slots_count: int) -> list[int]:
    if page_key == "page4_relationship":
        # Page4 is AI-only generation (no DB/Excel bank); use slot-index placeholders.
        return [i for i in range(1, slots_count + 1)]
    if page_key == "dragon_cinema":
        rows = conn.execute(
            "SELECT item_id FROM content_bank_rows "
            "WHERE page_key=? "
            "AND item_id NOT IN (SELECT item_id FROM content_items WHERE page_key=? AND used=1) "
            "AND item_id NOT IN (SELECT content_id FROM used_state WHERE page_key=? AND used=1) "
            "ORDER BY item_id ASC",
            (page_key, page_key, page_key),
        ).fetchall()
        ids = [int(r[0]) for r in rows if int(r[0]) > 0]
        out: list[int] = []
        ptr = 0
        for _ in range(slots_count):
            if ptr >= len(ids):
                out.append(0)
            else:
                out.append(ids[ptr])
                ptr += 1
        return out

    page_cfg = load_page_config(PROJECT_ROOT, page_key).profile
    content_cfg = page_cfg.get("content", {})
    provider = str(content_cfg.get("provider", "excel")).strip().lower()
    if provider == "db":
        rows = conn.execute(
            "SELECT item_id FROM content_bank_rows "
            "WHERE page_key=? "
            "AND item_id NOT IN (SELECT item_id FROM content_items WHERE page_key=? AND used=1) "
            "AND item_id NOT IN (SELECT content_id FROM used_state WHERE page_key=? AND used=1) "
            "ORDER BY item_id ASC",
            (page_key, page_key, page_key),
        ).fetchall()
        available_ids = [int(r[0]) for r in rows if int(r[0]) > 0]
        predicted: list[int] = []
        ptr = 0
        batch_size = int(content_cfg.get("batch_size", 1) or 1)
        for _ in range(slots_count):
            if ptr >= len(available_ids):
                predicted.append(0)
                continue
            predicted.append(available_ids[ptr])
            ptr += max(1, batch_size)
        return predicted

    xlsx_rel = str(content_cfg.get("xlsx_path", "")).strip()
    sheet_name = str(content_cfg.get("sheet_name", "")).strip()
    batch_size = int(content_cfg.get("batch_size", 1) or 1)
    if not xlsx_rel or not sheet_name:
        return [0] * slots_count
    xlsx_path = (PROJECT_ROOT / xlsx_rel).resolve()
    if not xlsx_path.exists():
        return [0] * slots_count

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    idx = {str(name): i + 1 for i, name in enumerate(header) if name is not None}
    id_col = idx.get("id") or idx.get("Number") or idx.get("No.")
    used_col = idx.get("used")
    if not id_col or not used_col:
        return [0] * slots_count

    # Preload used IDs once; per-row DB lookups are too slow for large sheets.
    used_ids: set[int] = set()
    for row in conn.execute("SELECT item_id, used FROM content_items WHERE page_key=?", (page_key,)).fetchall():
        try:
            if int(row[1] or 0) == 1:
                used_ids.add(int(row[0] or 0))
        except Exception:
            continue
    for row in conn.execute("SELECT content_id, used FROM used_state WHERE page_key=?", (page_key,)).fetchall():
        try:
            if int(row[1] or 0) == 1:
                used_ids.add(int(row[0] or 0))
        except Exception:
            continue

    available_ids: list[int] = []
    max_col = max(id_col, used_col)
    for row_vals in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
        try:
            item_id = int((row_vals[id_col - 1] if len(row_vals) >= id_col else 0) or 0)
        except Exception:
            item_id = 0
        if item_id <= 0:
            continue
        try:
            raw_used = row_vals[used_col - 1] if len(row_vals) >= used_col else 0
            excel_used = int(raw_used or 0) == 1
        except Exception:
            excel_used = False
        if excel_used or item_id in used_ids:
            continue
        available_ids.append(item_id)

    predicted: list[int] = []
    ptr = 0
    for _ in range(slots_count):
        if ptr >= len(available_ids):
            predicted.append(0)
            continue
        predicted.append(available_ids[ptr])
        ptr += max(1, batch_size)
    return predicted


def _upsert_quota(conn: sqlite3.Connection, page_key: str, *, state: str, backoff_minutes: int = 0, until: str = "", last_error: str = "") -> None:
    conn.execute(
        "INSERT INTO grok_quota_state(page_key,state,rate_limited_until,backoff_minutes,last_error,updated_at) VALUES(?,?,?,?,?,?) "
        "ON CONFLICT(page_key) DO UPDATE SET state=excluded.state,rate_limited_until=excluded.rate_limited_until,backoff_minutes=excluded.backoff_minutes,last_error=excluded.last_error,updated_at=excluded.updated_at",
        (page_key, state, until, backoff_minutes, last_error, _now_iso()),
    )
    conn.commit()


def _get_quota(conn: sqlite3.Connection, page_key: str) -> dict[str, Any]:
    row = conn.execute(
        "SELECT state,rate_limited_until,backoff_minutes,last_error FROM grok_quota_state WHERE page_key=?",
        (page_key,),
    ).fetchone()
    if not row:
        return {"state": "ok", "rate_limited_until": "", "backoff_minutes": 0, "last_error": ""}
    return {"state": row[0], "rate_limited_until": row[1], "backoff_minutes": int(row[2] or 0), "last_error": row[3] or ""}


def _increment_kpi(conn: sqlite3.Connection, run_date: str, page_key: str, **deltas: int) -> None:
    conn.execute(
        "INSERT INTO daily_kpi(run_date,page_key,attempted,generated,scheduled,failed,rate_limit_hits,retries,updated_at) VALUES(?,?,?,?,?,?,?,?,?) "
        "ON CONFLICT(run_date,page_key) DO NOTHING",
        (run_date, page_key, 0, 0, 0, 0, 0, 0, _now_iso()),
    )
    for k, v in deltas.items():
        conn.execute(f"UPDATE daily_kpi SET {k}={k}+?, updated_at=? WHERE run_date=? AND page_key=?", (int(v), _now_iso(), run_date, page_key))
    conn.commit()


def _reserve_content(conn: sqlite3.Connection, page_key: str, run_date: str, slot: str, content_id: int) -> str:
    rid = str(uuid.uuid4())
    conn.execute(
        "INSERT OR REPLACE INTO content_reservations(page_key,run_date,slot,content_id,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
        (page_key, run_date, slot, content_id, "reserved", _now_iso(), _now_iso()),
    )
    conn.commit()
    return rid


def _mark_reservation(conn: sqlite3.Connection, page_key: str, run_date: str, slot: str, status: str) -> None:
    conn.execute(
        "UPDATE content_reservations SET status=?, updated_at=? WHERE page_key=? AND run_date=? AND slot=?",
        (status, _now_iso(), page_key, run_date, slot),
    )
    conn.commit()


def _render_dashboard(report_path: Path) -> Path:
    payload = json.loads(report_path.read_text(encoding="utf-8"))
    out_dir = PROJECT_ROOT / "runs" / "reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    html_path = out_dir / f"dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    rows = []
    for r in payload.get("results", []):
        rows.append(f"<tr><td>{r.get('page_key')}</td><td>{r.get('slot')}</td><td>{r.get('state')}</td><td>{r.get('idempotency_key')}</td><td>{r.get('error','')}</td></tr>")
    html = f"""<!doctype html><html><head><meta charset='utf-8'><title>AutoVideoAgent Dashboard</title>
<style>body{{font-family:Arial;padding:16px}}table{{border-collapse:collapse;width:100%}}td,th{{border:1px solid #ccc;padding:8px}}</style></head>
<body><h2>Run Summary</h2>
<p>Rendered: {payload.get('totals',{}).get('rendered',0)} | Scheduled: {payload.get('totals',{}).get('scheduled',0)} | Failed: {payload.get('totals',{}).get('failed',0)} | Retries: {payload.get('totals',{}).get('retry_count',0)}</p>
<table><thead><tr><th>Page</th><th>Slot</th><th>State</th><th>Idempotency</th><th>Error</th></tr></thead><tbody>{''.join(rows)}</tbody></table>
</body></html>"""
    html_path.write_text(html, encoding="utf-8")
    return html_path


def _preflight(plans: list[PagePlan]) -> list[str]:
    issues: list[str] = []
    ffprobe = shutil.which("ffprobe")
    if not ffprobe:
        issues.append("ffprobe_not_found")
    # CDP endpoint check
    try:
        import urllib.request

        with urllib.request.urlopen("http://127.0.0.1:9100/json/version", timeout=5) as r:
            if r.status != 200:
                issues.append(f"cdp_unhealthy_status:{r.status}")
    except Exception as ex:
        issues.append(f"cdp_unreachable:{ex}")

    # DB writable check
    try:
        db = connect(PROJECT_ROOT / "data" / "v2" / "state.sqlite3")
        db.execute("CREATE TABLE IF NOT EXISTS __preflight_probe(k TEXT PRIMARY KEY, v TEXT)")
        db.execute("INSERT OR REPLACE INTO __preflight_probe(k,v) VALUES('last_check',?)", (_now_iso(),))
        db.commit()
    except Exception as ex:
        issues.append(f"db_not_writable:{ex}")

    for p in plans:
        try:
            if p.page_key == "dragon_cinema":
                # Dragon page uses custom generator; skip strict page schema validation.
                grok = Path.home() / ".grok" / "bin" / "grok.exe"
                if not grok.exists():
                    issues.append("grok_missing:dragon_cinema")
                continue
            cfg = load_page_config(PROJECT_ROOT, p.page_key).profile
            provider = str(cfg.get("content", {}).get("provider", "excel")).strip().lower()
            if provider != "db":
                xlsx = (PROJECT_ROOT / str(cfg.get("content", {}).get("xlsx_path", ""))).resolve()
                if not xlsx.exists():
                    issues.append(f"xlsx_missing:{p.page_key}")
            if _grok_dependent(p.page_key):
                grok = Path.home() / ".grok" / "bin" / "grok.exe"
                if not grok.exists():
                    issues.append(f"grok_missing:{p.page_key}")
        except Exception as ex:
            issues.append(f"config_error:{p.page_key}:{ex}")
    return issues


def _slot_priority(slots: list[str]) -> list[str]:
    def key(s: str) -> tuple[int, int]:
        hh, mm = [int(x) for x in s.split(":")]
        evening = 1 if hh >= 17 else 0
        return (evening, hh * 60 + mm)
    return sorted(slots, key=key, reverse=True)


def _timezone_guard() -> str:
    proc = subprocess.run(["powershell", "-NoProfile", "-Command", "(Get-TimeZone).Id"], capture_output=True, text=True)
    tz = (proc.stdout or "").strip() if proc.returncode == 0 else "unknown"
    return tz


def _within_cutoff(now: datetime, cutoff_hhmm: str) -> bool:
    hh, mm = [int(x) for x in cutoff_hhmm.split(":")]
    cutoff = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
    return now <= cutoff


def _schedule_dt_for_slot_next_day(slot_hhmm: str, now: datetime) -> datetime:
    hh, mm = [int(x) for x in slot_hhmm.split(":")]
    base = now + timedelta(days=1)
    return base.replace(hour=hh, minute=mm, second=0, microsecond=0)


def main() -> None:
    ap = argparse.ArgumentParser(description="Unified job runner with resilient Grok retry and scheduling.")
    ap.add_argument("--page", default="", help="Optional page key filter.")
    ap.add_argument("--render-profile", default="production", choices=["production", "fast_preview"])
    ap.add_argument("--catchup-today", action="store_true", help="Backfill missed same-day slots into next valid times.")
    args = ap.parse_args()

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = PROJECT_ROOT / "runs" / "daily_batch" / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    load_dotenv(PROJECT_ROOT / ".env")
    log = JsonLogger(run_dir / "events.jsonl")
    db = connect(PROJECT_ROOT / "data" / "v2" / "state.sqlite3")

    overrides = _read_json(PROJECT_ROOT / "control" / "manual_override.json", {"pause_pages": [], "skip_slots": {}})
    global_cfg = _read_json(PROJECT_ROOT / "configs" / "global.json", {})
    reliability = global_cfg.get("reliability", {}) if isinstance(global_cfg, dict) else {}
    cutoff_hhmm = str(reliability.get("daily_retry_cutoff", "23:30"))
    backoff_steps = reliability.get("grok_backoff_minutes", [30, 60, 120])
    min_daily_target = int(reliability.get("minimum_daily_target", 1))
    render_timeout_seconds = int(reliability.get("render_timeout_seconds", 900))

    killed = _cleanup_stale_processes(30)
    cleaned = _cleanup_old_runs(int(reliability.get("cleanup_days_to_keep", 7)))
    tz = _timezone_guard()
    log.write(page=args.page or "all", job_id="-", step="startup", status="ok", extra={"killed": killed, "cleaned_old_runs": cleaned, "timezone": tz})

    log.write(page=args.page or "all", job_id="-", step="load_plans", status="start")
    plans = _load_plans()
    log.write(page=args.page or "all", job_id="-", step="load_plans", status="ok", extra={"count": len(plans)})
    if args.page:
        plans = [p for p in plans if p.page_key == args.page]
    plans = [p for p in plans if p.page_key not in set(overrides.get("pause_pages", []))]
    if not plans:
        raise RuntimeError("No plans selected.")

    log.write(page=args.page or "all", job_id="-", step="preflight", status="start")
    issues = _preflight(plans)
    log.write(page=args.page or "all", job_id="-", step="preflight", status="ok", extra={"issues": len(issues)})
    for i in issues:
        log.write(page="all", job_id="-", step="preflight", status="warn", error=i)

    now = datetime.now()
    pending: list[PendingItem] = []
    backfill_events: list[dict[str, str]] = []
    log.write(page=args.page or "all", job_id="-", step="queue_build", status="start")
    for plan in plans:
        run_date = now.strftime("%Y-%m-%d")
        slots = [s for s in _slot_priority(plan.slots) if s not in set(overrides.get("skip_slots", {}).get(plan.page_key, []))]
        predicted_ids = _predict_next_content_ids(db, plan.page_key, len(slots))
        backfill_cursor = now + timedelta(minutes=25)
        for slot in slots:
            predicted_content_id = predicted_ids.pop(0) if predicted_ids else 0
            idem = _build_idempotency_key(plan.page_key, run_date, slot, predicted_content_id)
            if predicted_content_id > 0 and _job_exists_by_idem(db, idem):
                continue
            target_dt = _schedule_dt_for_slot_next_day(slot, now)
            # If catchup mode is on and the wall-clock slot for *today* is already past,
            # remap it into the next valid future window instead of letting it roll to tomorrow.
            hh, mm = [int(x) for x in slot.split(":")]
            slot_today = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
            slot_missed_today = slot_today < now
            if args.catchup_today and slot_missed_today:
                mapped = max(backfill_cursor, now + timedelta(minutes=25))
                backfill_cursor = mapped + timedelta(minutes=30)
                log.write(
                    page=plan.page_key,
                    job_id="-",
                    step="queue_build",
                    status="info",
                    error=f"MISSED_SLOT slot={slot} remapped_to={mapped.strftime('%H:%M')}",
                )
                backfill_events.append(
                    {
                        "page_key": plan.page_key,
                        "original_slot": slot,
                        "mapped_slot": mapped.strftime("%H:%M"),
                    }
                )
                target_dt = mapped

            if target_dt <= now + timedelta(minutes=20):
                if args.catchup_today:
                    mapped = max(backfill_cursor, now + timedelta(minutes=25))
                    backfill_cursor = mapped + timedelta(minutes=30)
                    log.write(
                        page=plan.page_key,
                        job_id="-",
                        step="queue_build",
                        status="info",
                        error=f"MISSED_SLOT slot={slot} remapped_to={mapped.strftime('%H:%M')}",
                    )
                    backfill_events.append(
                        {
                            "page_key": plan.page_key,
                            "original_slot": slot,
                            "mapped_slot": mapped.strftime("%H:%M"),
                        }
                    )
                    target_dt = mapped
                else:
                    log.write(
                        page=plan.page_key,
                        job_id="-",
                        step="queue_build",
                        status="info",
                        error=f"MISSED_SLOT slot={slot} skipped_past_window",
                    )
                    continue
            if target_dt >= now + timedelta(days=29):
                continue
            rid = _reserve_content(db, plan.page_key, run_date, slot, predicted_content_id)
            pending.append(PendingItem(plan.page_key, slot, run_date, target_dt, plan.asset_id, plan.generator_type, predicted_content_id, rid))

    queue_path = run_dir / "daily_batch_queue.json"
    queue_path.write_text(json.dumps({"run_id": run_id, "items": [p.__dict__ for p in pending]}, indent=2, default=str), encoding="utf-8")
    log.write(page=args.page or "all", job_id="-", step="queue_build", status="ok", extra={"pending": len(pending)})

    plan_map = {p.page_key: p for p in plans}
    results: list[dict[str, Any]] = []
    retry_count = 0

    log.write(page=args.page or "all", job_id="-", step="browser_connect", status="start")
    with sync_playwright() as pw:
        browser = pw.chromium.connect_over_cdp("http://127.0.0.1:9100", timeout=120000)
        ctx = browser.contexts[0] if browser.contexts else browser.new_context()
        shots = run_dir / "screens"
        shots.mkdir(parents=True, exist_ok=True)
        log.write(page=args.page or "all", job_id="-", step="browser_connect", status="ok", extra={"contexts": len(browser.contexts)})

        while pending and _within_cutoff(datetime.now(), cutoff_hhmm):
            progressed = False
            for item in list(pending):
                page_key = item.page_key
                quota = _get_quota(db, page_key)
                if quota["state"] == "rate_limited_until" and quota["rate_limited_until"]:
                    try:
                        if datetime.now() < datetime.fromisoformat(str(quota["rate_limited_until"])):
                            continue
                    except Exception:
                        pass

                job_id = str(uuid.uuid4())
                _insert_job(db, job_id=job_id, page_key=page_key, run_date=item.run_date, slot=item.slot)
                _state_set(db, job_id, "rendering")
                _increment_kpi(db, item.run_date, page_key, attempted=1)
                log.write(page=page_key, job_id=job_id, step="rendering", status="start", extra={"slot": item.slot})

                try:
                    t0 = time.time()
                    manifest, payload, caption = _run_render(
                        page_key,
                        heartbeat=lambda sec: log.write(
                            page=page_key,
                            job_id=job_id,
                            step="rendering",
                            status="in_progress",
                            extra={"slot": item.slot, "elapsed_sec": sec},
                        ),
                        timeout_seconds=render_timeout_seconds,
                        slot=item.slot,
                        target_dt=item.target_dt,
                    )
                    elapsed = round(time.time() - t0, 2)
                    log.write(page=page_key, job_id=job_id, step="rendering", status="ok", extra={"slot": item.slot, "seconds": elapsed})
                    video = Path(payload["output_mp4"])
                    _state_set(db, job_id, "rendered", video_path=str(video))
                    points = (payload.get("spec") or {}).get("points") or []
                    content_id = int((points[0] or {}).get("source_item_id", 0)) if points else item.predicted_content_id
                    idem = _build_idempotency_key(page_key, item.run_date, item.slot, content_id)
                    if _job_exists_by_idem(db, idem):
                        _state_set(db, job_id, "failed", error=f"duplicate_idempotency:{idem}")
                        _mark_reservation(db, page_key, item.run_date, item.slot, "duplicate")
                        pending.remove(item)
                        continue
                    duration, has_audio = _probe_video(video)
                    require_audio = page_key != "dragon_cinema"
                    if (not video.exists()) or duration <= 0.5 or (require_audio and (not has_audio)) or (not caption.strip()):
                        raise RuntimeError(f"health_check_failed:{video.exists()}:{duration}:{has_audio}:{bool(caption.strip())}")

                    _state_set(db, job_id, "uploading", content_id=content_id, idempotency_key=idem, manifest_path=str(manifest), video_path=str(video), caption=caption, target_time=item.target_dt.isoformat())
                    log.write(
                        page=page_key,
                        job_id=job_id,
                        step="uploading",
                        status="start",
                        extra={"slot": item.slot, "video": str(video), "target": item.target_dt.isoformat(timespec="minutes")},
                    )

                    status = _upload_and_schedule_with_retry(
                        ctx,
                        PagePlan(page_key, item.asset_id, item.generator_type, []),
                        str(video),
                        caption,
                        item.target_dt,
                        shots,
                    )
                    if status.get("retried"):
                        retry_count += 1
                        _increment_kpi(db, item.run_date, page_key, retries=1)
                        log.write(
                            page=page_key,
                            job_id=job_id,
                            step="uploading",
                            status="retry",
                            extra={"slot": item.slot},
                        )
                    if status.get("ok"):
                        log.write(
                            page=page_key,
                            job_id=job_id,
                            step="uploading",
                            status="ok",
                            extra={"slot": item.slot},
                        )
                        _state_set(db, job_id, "scheduled")
                        _mark_reservation(db, page_key, item.run_date, item.slot, "scheduled")
                        _increment_kpi(db, item.run_date, page_key, generated=1, scheduled=1)
                        _upsert_quota(db, page_key, state="ok", backoff_minutes=0, until="", last_error="")
                        results.append({"job_id": job_id, "page_key": page_key, "slot": item.slot, "idempotency_key": idem, "state": "scheduled", "error": ""})
                        log.write(
                            page=page_key,
                            job_id=job_id,
                            step="scheduled",
                            status="ok",
                            extra={"slot": item.slot, "idempotency_key": idem},
                        )
                        print(f"SCHEDULED_OK page={page_key} slot={item.slot} idem={idem}")
                        send_telegram(
                            reel_status_message(
                                page_key=page_key,
                                slot=item.slot,
                                scheduled_for=item.target_dt.isoformat(timespec="minutes"),
                                status="scheduled",
                                video=str(video),
                            )
                        )
                        pending.remove(item)
                        progressed = True
                    else:
                        log.write(
                            page=page_key,
                            job_id=job_id,
                            step="uploading",
                            status="failed",
                            error=json.dumps(status),
                            extra={"slot": item.slot},
                        )
                        _state_set(db, job_id, "failed", error=json.dumps(status))
                        _increment_kpi(db, item.run_date, page_key, failed=1)
                        results.append({"job_id": job_id, "page_key": page_key, "slot": item.slot, "idempotency_key": idem, "state": "failed", "error": json.dumps(status)})
                        send_telegram(
                            reel_status_message(
                                page_key=page_key,
                                slot=item.slot,
                                scheduled_for=item.target_dt.isoformat(timespec="minutes"),
                                status="failed_schedule",
                                video=str(video),
                                error=json.dumps(status),
                            )
                        )
                        pending.remove(item)
                        progressed = True
                except Exception as ex:
                    err = f"{type(ex).__name__}: {ex}"
                    klass = _classify_generation_error(err)
                    if _grok_dependent(page_key) and klass == "grok_rate_limit":
                        b = int(quota.get("backoff_minutes", 0) or 0)
                        if b <= 0:
                            b = int(backoff_steps[0])
                        else:
                            nxt = [x for x in backoff_steps if x > b]
                            b = int(nxt[0] if nxt else backoff_steps[-1])
                        until = (datetime.now() + timedelta(minutes=b)).isoformat(timespec="seconds")
                        _upsert_quota(db, page_key, state="rate_limited_until", backoff_minutes=b, until=until, last_error=err)
                        _increment_kpi(db, item.run_date, page_key, rate_limit_hits=1)
                        _state_set(db, job_id, "retry_wait", error=err)
                        log.write(page=page_key, job_id=job_id, step="grok_cooldown", status="wait", extra={"minutes": b, "until": until})
                        send_telegram(
                            reel_status_message(
                                page_key=page_key,
                                slot=item.slot,
                                scheduled_for=item.target_dt.isoformat(timespec="minutes"),
                                status="grok_limit_exceeded",
                                error=f"{err[:220]} | cooldown_until={until} | retry_in_minutes={b}",
                            )
                        )
                        continue
                    _state_set(db, job_id, "failed", error=err)
                    _increment_kpi(db, item.run_date, page_key, failed=1)
                    _mark_reservation(db, page_key, item.run_date, item.slot, "failed")
                    results.append({"job_id": job_id, "page_key": page_key, "slot": item.slot, "idempotency_key": "", "state": "failed", "error": err})
                    send_telegram(
                        reel_status_message(
                            page_key=page_key,
                            slot=item.slot,
                            scheduled_for=item.target_dt.isoformat(timespec="minutes"),
                            status="failed_render_or_upload",
                            error=err,
                        )
                    )
                    pending.remove(item)
                    progressed = True

            if not pending:
                break
            if not progressed:
                next_wait = 30
                for p in {x.page_key for x in pending}:
                    q = _get_quota(db, p)
                    try:
                        if q["state"] == "rate_limited_until" and q["rate_limited_until"]:
                            dt = datetime.fromisoformat(q["rate_limited_until"])
                            mins = max(1, int((dt - datetime.now()).total_seconds() // 60))
                            next_wait = min(next_wait, mins)
                    except Exception:
                        continue
                wait_minutes = max(1, min(next_wait, 30))
                log.write(page="all", job_id="-", step="idle_wait", status="wait", extra={"minutes": wait_minutes, "pending": len(pending)})
                subprocess.run(["powershell", "-NoProfile", "-Command", f"Start-Sleep -Seconds {wait_minutes * 60}"], check=False)

    # minimum target informational check
    run_date = datetime.now().strftime("%Y-%m-%d")
    for p in plans:
        row = db.execute("SELECT scheduled FROM daily_kpi WHERE run_date=? AND page_key=?", (run_date, p.page_key)).fetchone()
        scheduled = int(row[0]) if row else 0
        if scheduled < min_daily_target:
            log.write(page=p.page_key, job_id="-", step="minimum_target", status="warn", error=f"scheduled={scheduled} below target={min_daily_target}")

    totals = {
        "rendered": sum(1 for r in results if r["state"] in {"scheduled", "failed"}),
        "posted": 0,
        "scheduled": sum(1 for r in results if r["state"] == "scheduled"),
        "failed": sum(1 for r in results if r["state"] == "failed"),
        "retry_count": retry_count,
        "pending_after_cutoff": len(pending),
    }
    report = {"run_id": run_id, "queue_file": str(queue_path), "totals": totals, "results": results, "pending": [p.__dict__ for p in pending]}
    if backfill_events:
        report["backfilled_slots"] = backfill_events
    report_path = run_dir / "summary_report.json"
    report_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    dashboard = _render_dashboard(report_path)
    msg = (
        f"AutoVideoAgent KPI\\n"
        f"run_id: {run_id}\\n"
        f"scheduled: {totals['scheduled']}\\n"
        f"failed: {totals['failed']}\\n"
        f"pending: {totals['pending_after_cutoff']}\\n"
        f"retries: {totals['retry_count']}\\n"
        f"report: {report_path}"
    )
    send_telegram(msg)
    log.write(
        page=args.page or "all",
        job_id="-",
        step="run_summary",
        status="ok",
        extra={"scheduled": totals["scheduled"], "failed": totals["failed"], "pending": totals["pending_after_cutoff"], "retries": totals["retry_count"]},
    )
    print(
        "RUN_SUMMARY "
        f"scheduled={totals['scheduled']} failed={totals['failed']} "
        f"pending={totals['pending_after_cutoff']} retries={totals['retry_count']}"
    )
    print(f"RUN_DIR={run_dir}")
    print(f"QUEUE={queue_path}")
    print(f"REPORT={report_path}")
    print(f"DASHBOARD={dashboard}")
    print(f"SUCCESS={totals['scheduled']}/{max(1, len(results))}")


if __name__ == "__main__":
    main()

from __future__ import annotations

import argparse
import json
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _cell(v: Any, default: Any = "") -> Any:
    return default if v is None else v


def _to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "y", "on"}


def _to_time_hhmm(v: Any, default: str = "09:00") -> str:
    s = str(_cell(v, default)).strip()
    if len(s) >= 5 and s[2] == ":":
        return s[:5]
    return default


def build_runtime_from_excel(xlsx_path: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx_path)
    ws_pages = wb["pages"]
    ws_slots = wb["posting_slots"]
    ws_jobs = wb["agent_jobs"]

    pages_header = [c.value for c in ws_pages[1]]
    slots_header = [c.value for c in ws_slots[1]]
    jobs_header = [c.value for c in ws_jobs[1]]
    pidx = {name: i + 1 for i, name in enumerate(pages_header)}
    sidx = {name: i + 1 for i, name in enumerate(slots_header)}
    jidx = {name: i + 1 for i, name in enumerate(jobs_header)}

    pages: list[dict[str, Any]] = []
    for r in range(2, ws_pages.max_row + 1):
        page_key = str(_cell(ws_pages.cell(r, pidx["page_key"]).value)).strip()
        if not page_key:
            continue
        enabled = _to_bool(ws_pages.cell(r, pidx["enabled"]).value)
        slots: list[str] = []
        for sr in range(2, ws_slots.max_row + 1):
            if str(_cell(ws_slots.cell(sr, sidx["page_key"]).value)).strip() != page_key:
                continue
            if not _to_bool(ws_slots.cell(sr, sidx["active"]).value):
                continue
            slots.append(_to_time_hhmm(ws_slots.cell(sr, sidx["post_time"]).value))
        pages.append(
            {
                "page_key": page_key,
                "enabled": enabled,
                "mode": str(_cell(ws_pages.cell(r, pidx["mode"]).value)).strip() or "ui_schedule",
                "facebook_asset_id": str(_cell(ws_pages.cell(r, pidx["facebook_asset_id"]).value)).strip(),
                "generator_type": str(_cell(ws_pages.cell(r, pidx["generator_type"]).value)).strip(),
                "content_source_file": str(_cell(ws_pages.cell(r, pidx["content_source_file"]).value)).strip(),
                "logo_file": str(_cell(ws_pages.cell(r, pidx["logo_file"]).value)).strip(),
                "timezone": str(_cell(ws_pages.cell(r, pidx["timezone"]).value, "Asia/Calcutta")).strip(),
                "morning_run_time": _to_time_hhmm(ws_pages.cell(r, pidx["morning_run_time"]).value),
                "schedule_horizon_days": int(_cell(ws_pages.cell(r, pidx["schedule_horizon_days"]).value, 1)),
                "posting_slots": sorted(slots),
            }
        )

    job = {
        "job_key": "daily_ui_batch",
        "enabled": True,
        "run_time": "09:00",
        "run_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
        "scope": "all_enabled_pages",
        "action": "generate_and_schedule",
    }
    for r in range(2, ws_jobs.max_row + 1):
        if str(_cell(ws_jobs.cell(r, jidx["job_key"]).value)).strip() != "daily_ui_batch":
            continue
        job = {
            "job_key": "daily_ui_batch",
            "enabled": _to_bool(ws_jobs.cell(r, jidx["enabled"]).value),
            "run_time": _to_time_hhmm(ws_jobs.cell(r, jidx["run_time"]).value, "09:00"),
            "run_days": [
                x.strip()
                for x in str(_cell(ws_jobs.cell(r, jidx["run_days"]).value, "Mon,Tue,Wed,Thu,Fri,Sat,Sun")).split(",")
                if x.strip()
            ],
            "scope": str(_cell(ws_jobs.cell(r, jidx["scope"]).value, "all_enabled_pages")).strip(),
            "action": str(_cell(ws_jobs.cell(r, jidx["action"]).value, "generate_and_schedule")).strip(),
        }
        break

    return {"generated_from": str(xlsx_path), "pages": pages, "job": job}


def update_task_schedule(task_name: str, run_time: str, enabled: bool) -> None:
    subprocess.run(["schtasks", "/Change", "/TN", task_name, "/ST", run_time], check=True, capture_output=True, text=True)
    if enabled:
        subprocess.run(["schtasks", "/Change", "/TN", task_name, "/ENABLE"], check=True, capture_output=True, text=True)
    else:
        subprocess.run(["schtasks", "/Change", "/TN", task_name, "/DISABLE"], check=True, capture_output=True, text=True)


def main() -> None:
    ap = argparse.ArgumentParser(description="Sync control Excel into runtime JSON and Task Scheduler.")
    ap.add_argument("--project-root", default=str(Path(__file__).resolve().parents[1]))
    ap.add_argument("--task-name", default="\\AutoVideoAgent_Daily_UI_Batch_9AM")
    args = ap.parse_args()

    root = Path(args.project_root).resolve()
    xlsx = root / "control" / "page_automation_control.xlsx"
    runtime_json = root / "control" / "automation_runtime.json"

    if not xlsx.exists():
        raise FileNotFoundError(f"Control Excel not found: {xlsx}")

    runtime = build_runtime_from_excel(xlsx)
    runtime_json.write_text(json.dumps(runtime, ensure_ascii=False, indent=2), encoding="utf-8")

    job = runtime["job"]
    update_task_schedule(args.task_name, job["run_time"], bool(job["enabled"]))

    print(f"CONTROL_XLSX={xlsx}")
    print(f"RUNTIME_JSON={runtime_json}")
    print(f"TASK_NAME={args.task_name}")
    print(f"TASK_RUN_TIME={job['run_time']}")
    print(f"TASK_ENABLED={job['enabled']}")


if __name__ == "__main__":
    main()

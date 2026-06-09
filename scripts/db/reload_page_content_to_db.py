from __future__ import annotations

import argparse
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from autovideo.services.state_store import connect  # noqa: E402


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _clear_page_content(conn: sqlite3.Connection, page_key: str, reset_used: bool) -> None:
    conn.execute("DELETE FROM content_bank_rows WHERE page_key=?", (page_key,))
    conn.execute("DELETE FROM content_rows WHERE page_key=?", (page_key,))
    conn.execute("DELETE FROM captions WHERE page_key=?", (page_key,))
    conn.execute("DELETE FROM hashtags WHERE page_key=?", (page_key,))
    conn.execute("DELETE FROM scene_prompts WHERE page_key=?", (page_key,))
    if reset_used:
        conn.execute("DELETE FROM content_items WHERE page_key=?", (page_key,))
        conn.execute("DELETE FROM used_state WHERE page_key=?", (page_key,))
    conn.commit()


def _load_female(conn: sqlite3.Connection, page_key: str, xlsx: Path, sheet: str) -> int:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    idx = {str(n): i + 1 for i, n in enumerate(hdr) if n is not None}
    required = ["id", "heading_line1", "heading_line2", "point_text", "highlight_first_words", "cta", "used"]
    for k in required:
        if k not in idx:
            raise ValueError(f"Missing column '{k}' in {xlsx}")

    rows = []
    used_rows = []
    ts = _now()
    for r in range(2, ws.max_row + 1):
        item_id = int(ws.cell(r, idx["id"]).value or 0)
        if item_id <= 0:
            continue
        h1 = str(ws.cell(r, idx["heading_line1"]).value or "").strip()
        h2 = str(ws.cell(r, idx["heading_line2"]).value or "").strip()
        pt = str(ws.cell(r, idx["point_text"]).value or "").strip()
        hfw = int(ws.cell(r, idx["highlight_first_words"]).value or 3)
        cta = str(ws.cell(r, idx["cta"]).value or "").strip()
        used = int(ws.cell(r, idx["used"]).value or 0) == 1
        rows.append((page_key, item_id, h1, h2, pt, hfw, cta, "", "", "", ts))
        if used:
            used_rows.append((page_key, item_id, 1, ts))

    conn.executemany(
        "INSERT INTO content_bank_rows("
        "page_key,item_id,heading_line1,heading_line2,point_text,highlight_first_words,cta,scene_prompt,caption,hashtags,updated_at"
        ") VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    if used_rows:
        conn.executemany(
            "INSERT INTO content_items(page_key,item_id,used,used_at) VALUES(?,?,?,?) "
            "ON CONFLICT(page_key,item_id) DO UPDATE SET used=excluded.used, used_at=excluded.used_at",
            used_rows,
        )
        conn.executemany(
            "INSERT INTO used_state(page_key,content_id,used,used_at) VALUES(?,?,?,?) "
            "ON CONFLICT(page_key,content_id) DO UPDATE SET used=excluded.used, used_at=excluded.used_at",
            used_rows,
        )
    conn.commit()
    return len(rows)


def _load_desire(conn: sqlite3.Connection, page_key: str, xlsx: Path, sheet: str) -> int:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    idx = {str(n): i + 1 for i, n in enumerate(hdr) if n is not None}
    required = ["Number", "Heading", "Hook", "Reason", "Caption", "Hashtags", "scene_prompt", "used"]
    for k in required:
        if k not in idx:
            raise ValueError(f"Missing column '{k}' in {xlsx}")

    rows = []
    used_rows = []
    ts = _now()
    for r in range(2, ws.max_row + 1):
        item_id = int(ws.cell(r, idx["Number"]).value or 0)
        if item_id <= 0:
            continue
        heading = str(ws.cell(r, idx["Heading"]).value or "").strip()
        hook = str(ws.cell(r, idx["Hook"]).value or "").strip()
        reason = str(ws.cell(r, idx["Reason"]).value or "").strip()
        caption = str(ws.cell(r, idx["Caption"]).value or "").strip()
        hashtags = str(ws.cell(r, idx["Hashtags"]).value or "").strip()
        scene_prompt = str(ws.cell(r, idx["scene_prompt"]).value or "").strip()
        used = int(ws.cell(r, idx["used"]).value or 0) == 1
        point_text = f"{hook}||{reason}" if reason else hook
        rows.append((page_key, item_id, "DAILY DESIRE FACTS", heading, point_text, 3, "", scene_prompt, caption, hashtags, ts))
        if used:
            used_rows.append((page_key, item_id, 1, ts))

    conn.executemany(
        "INSERT INTO content_bank_rows("
        "page_key,item_id,heading_line1,heading_line2,point_text,highlight_first_words,cta,scene_prompt,caption,hashtags,updated_at"
        ") VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    if used_rows:
        conn.executemany(
            "INSERT INTO content_items(page_key,item_id,used,used_at) VALUES(?,?,?,?) "
            "ON CONFLICT(page_key,item_id) DO UPDATE SET used=excluded.used, used_at=excluded.used_at",
            used_rows,
        )
        conn.executemany(
            "INSERT INTO used_state(page_key,content_id,used,used_at) VALUES(?,?,?,?) "
            "ON CONFLICT(page_key,content_id) DO UPDATE SET used=excluded.used, used_at=excluded.used_at",
            used_rows,
        )
    conn.commit()
    return len(rows)


def _load_dragon(conn: sqlite3.Connection, page_key: str, xlsx: Path, sheet: str) -> int:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    idx = {str(n): i + 1 for i, n in enumerate(hdr) if n is not None}
    required = ["No.", "Scene 1", "Scene 2", "Caption", "Hashtags", "used"]
    for k in required:
        if k not in idx:
            raise ValueError(f"Missing column '{k}' in {xlsx}")

    rows = []
    used_rows = []
    ts = _now()
    for r in range(2, ws.max_row + 1):
        item_id = int(ws.cell(r, idx["No."]).value or 0)
        if item_id <= 0:
            continue
        s1 = str(ws.cell(r, idx["Scene 1"]).value or "").strip()
        s2 = str(ws.cell(r, idx["Scene 2"]).value or "").strip()
        caption = str(ws.cell(r, idx["Caption"]).value or "").strip()
        hashtags = str(ws.cell(r, idx["Hashtags"]).value or "").strip()
        used = int(ws.cell(r, idx["used"]).value or 0) == 1
        point_text = f"{s1}||{s2}" if s2 else s1
        rows.append((page_key, item_id, "FLORA KNOWS", "NOTHING", point_text, 3, "", s1, caption, hashtags, ts))
        if used:
            used_rows.append((page_key, item_id, 1, ts))

    conn.executemany(
        "INSERT INTO content_bank_rows("
        "page_key,item_id,heading_line1,heading_line2,point_text,highlight_first_words,cta,scene_prompt,caption,hashtags,updated_at"
        ") VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    if used_rows:
        conn.executemany(
            "INSERT INTO content_items(page_key,item_id,used,used_at) VALUES(?,?,?,?) "
            "ON CONFLICT(page_key,item_id) DO UPDATE SET used=excluded.used, used_at=excluded.used_at",
            used_rows,
        )
        conn.executemany(
            "INSERT INTO used_state(page_key,content_id,used,used_at) VALUES(?,?,?,?) "
            "ON CONFLICT(page_key,content_id) DO UPDATE SET used=excluded.used, used_at=excluded.used_at",
            used_rows,
        )
    conn.commit()
    return len(rows)


def _reload_to_db(db_path: Path, page_key: str, xlsx: Path, sheet: str, reset_used: bool) -> int:
    conn = connect(db_path)
    _clear_page_content(conn, page_key, reset_used=reset_used)
    if page_key == "female_psychology":
        return _load_female(conn, page_key, xlsx, sheet)
    if page_key == "daily_desire_facts":
        return _load_desire(conn, page_key, xlsx, sheet)
    if page_key == "dragon_cinema":
        return _load_dragon(conn, page_key, xlsx, sheet)
    raise ValueError(f"Unsupported page_key: {page_key}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Overwrite-load page Excel into SQLite DB(s).")
    ap.add_argument("--page", required=True, choices=["female_psychology", "daily_desire_facts", "dragon_cinema"])
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--reset-used", action="store_true")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.is_absolute():
        xlsx = (PROJECT_ROOT / xlsx).resolve()

    db_targets = [
        PROJECT_ROOT / "pages" / args.page / "data" / "state.sqlite3",
        PROJECT_ROOT / "data" / "v2" / "state.sqlite3",
    ]

    for db_path in db_targets:
        count = _reload_to_db(db_path=db_path, page_key=args.page, xlsx=xlsx, sheet=args.sheet, reset_used=args.reset_used)
        print(f"PAGE={args.page}")
        print(f"DB={db_path}")
        print(f"ROWS_LOADED={count}")
        print(f"RESET_USED={args.reset_used}")


if __name__ == "__main__":
    main()


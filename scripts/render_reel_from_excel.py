import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from render_reel_from_json import render_cfg


def _cell(v, default=""):
    return default if v is None else v


def load_next_five_rows(xlsx_path: Path, sheet_name: str) -> tuple[list[int], list[int], dict]:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    required = ["id", "heading_line1", "heading_line2", "point_text", "highlight_first_words", "cta", "used"]
    for k in required:
        if k not in idx:
            raise ValueError(f"Missing required column '{k}' in {xlsx_path}")

    picked_rows = []
    picked_ids = []
    for r in range(2, ws.max_row + 1):
        used = _cell(ws.cell(r, idx["used"] + 1).value, 0)
        if int(used) != 0:
            continue
        picked_rows.append(r)
        picked_ids.append(int(_cell(ws.cell(r, idx["id"] + 1).value, 0)))
        if len(picked_rows) == 5:
            break

    if len(picked_rows) < 5:
        raise ValueError("Not enough unused rows. Need at least 5.")

    first_row = picked_rows[0]
    heading = [
        str(_cell(ws.cell(first_row, idx["heading_line1"] + 1).value)).strip(),
        str(_cell(ws.cell(first_row, idx["heading_line2"] + 1).value)).strip(),
    ]
    cta = str(_cell(ws.cell(first_row, idx["cta"] + 1).value)).strip()

    points = []
    for r in picked_rows:
        text = str(_cell(ws.cell(r, idx["point_text"] + 1).value)).strip()
        hi = int(_cell(ws.cell(r, idx["highlight_first_words"] + 1).value, 3))
        points.append({"text": text, "highlight_first_words": hi})
    content = {"heading": heading, "points": points, "cta": cta}
    return picked_ids, picked_rows, content


def mark_rows_used(xlsx_path: Path, sheet_name: str, row_numbers: list[int]) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    for r in row_numbers:
        ws.cell(r, idx["used"] + 1, 1)
    wb.save(xlsx_path)


def main(xlsx_path: Path, style_json: Path, out_dir: Path, sheet_name: str, audio_override: str | None) -> None:
    style_cfg = json.loads(style_json.read_text(encoding="utf-8-sig"))
    picked_ids, picked_rows, content = load_next_five_rows(xlsx_path, sheet_name)
    cfg = dict(style_cfg)
    cfg["content"] = content
    if audio_override:
        cfg["assets"]["audio_path"] = audio_override

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = f"reel_from_excel_{stamp}_rows_{picked_ids[0]}_{picked_ids[-1]}"
    ass_path, mp4_path, frame_path = render_cfg(cfg, stem, out_dir)
    mark_rows_used(xlsx_path, sheet_name, picked_rows)
    print(f"USED_IDS={picked_ids}")
    print(ass_path)
    print(mp4_path)
    print(frame_path)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="pages/page1_female_psychology/content/reel_content_bank.xlsx")
    ap.add_argument("--style-json", default="output/exact_clone/reel_female_psychology_v2.json")
    ap.add_argument("--out-dir", default="output/exact_clone")
    ap.add_argument("--sheet", default="content_pool")
    ap.add_argument("--audio", default="assets/music/bg.mp3")
    args = ap.parse_args()
    main(Path(args.xlsx), Path(args.style_json), Path(args.out_dir), args.sheet, args.audio)



from __future__ import annotations

import random
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from autovideo.services.config_loader import load_page_config


DEFAULT_HASHTAG_POOL = [
    "#femalepsychology",
    "#datingtips",
    "#relationshipadvice",
    "#womendating",
    "#attraction",
    "#selfrespect",
    "#emotionalintelligence",
    "#loveadvice",
    "#modernromance",
    "#healthyrelationships",
    "#confidence",
    "#communication",
    "#datingadviceforwomen",
    "#highvaluedating",
    "#relationshipgoals",
    "#selflovejourney",
    "#feminineenergy",
    "#mindsetshift",
    "#personalgrowth",
    "#boundaries",
    "#healingjourney",
    "#secureattachment",
    "#relationshipcoach",
    "#usa",
    "#unitedstates",
    "#americanwomen",
    "#usdating",
    "#newyork",
    "#california",
    "#texas",
    "#florida",
    "#chicago",
    "#losangeles",
    "#miami",
    "#dallas",
    "#houston",
    "#atlanta",
    "#seattle",
    "#boston",
    "#washingtondc",
    "#sandiego",
    "#phoenix",
    "#lasvegas",
    "#philadelphia",
    "#charlotte",
    "#denver",
    "#austin",
]

PAGE_HASHTAG_POOLS = {
    "daily_desire_facts": [
        "#dailydesirefacts",
        "#relationshipfacts",
        "#attractionfacts",
        "#datingfacts",
        "#lovefacts",
        "#girlfacts",
        "#guyfacts",
        "#psychologyfacts",
        "#modernrelationships",
        "#emotionalattraction",
        "#confidence",
        "#communication",
        "#datingtips",
        "#relationshipadvice",
        "#usa",
        "#unitedstates",
        "#newyork",
        "#california",
        "#texas",
        "#florida",
        "#chicago",
        "#losangeles",
        "#miami",
        "#dallas",
        "#houston",
    ],
}


def _content_xlsx(project_root: Path, page_key: str) -> Path | None:
    try:
        cfg = load_page_config(project_root, page_key).profile
    except Exception:
        return None
    rel = str(cfg.get("content", {}).get("xlsx_path", "")).strip()
    if not rel:
        return None
    p = Path(rel)
    return p if p.is_absolute() else (project_root / p)


def _load_hashtag_pool(project_root: Path, page_key: str) -> list[str]:
    candidates: list[Path] = []
    xlsx = _content_xlsx(project_root, page_key)
    if xlsx:
        candidates.append(xlsx.parent / "high_rpm_hashtags.txt")
    candidates.append(project_root / "pages" / page_key / "content" / "high_rpm_hashtags.txt")

    for tag_file in candidates:
        if tag_file.exists():
            lines = [x.strip() for x in tag_file.read_text(encoding="utf-8").splitlines()]
            tags = [x if x.startswith("#") else f"#{x}" for x in lines if x and not x.startswith("//")]
            if tags:
                return list(dict.fromkeys(tags))
    return PAGE_HASHTAG_POOLS.get(page_key, DEFAULT_HASHTAG_POOL)


def _caption_from_excel(manifest: dict[str, Any], project_root: Path) -> str:
    spec = manifest.get("spec", {})
    page_key = str(spec.get("page_key", "")).strip().lower()
    if page_key != "daily_desire_facts":
        return ""

    points = spec.get("points", [])
    if not points:
        return ""
    row_id = points[0].get("source_item_id")
    if row_id is None:
        return ""

    xlsx = _content_xlsx(project_root, page_key)
    if not xlsx or not xlsx.exists():
        return ""

    wb = load_workbook(xlsx, data_only=True)
    sheet = wb["content_pool"] if "content_pool" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(sheet.cell(1, c).value or "").strip().lower() for c in range(1, sheet.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}
    num_col = idx.get("number")
    cap_col = idx.get("caption")
    tag_col = idx.get("hashtags")
    if not num_col or not cap_col:
        return ""

    caption = ""
    hashtags = ""
    for r in range(2, sheet.max_row + 1):
        val = sheet.cell(r, num_col).value
        try:
            current_id = int(val)
        except Exception:
            continue
        if current_id == int(row_id):
            caption = str(sheet.cell(r, cap_col).value or "").strip()
            if tag_col:
                hashtags = str(sheet.cell(r, tag_col).value or "").strip()
            break

    if not caption:
        return ""
    return f"{caption}\n\n{hashtags}".strip() if hashtags else caption


def build_caption(manifest: dict[str, Any], project_root: Path, hashtag_count: int = 5) -> str:
    excel_caption = _caption_from_excel(manifest, project_root)
    if excel_caption:
        return excel_caption

    spec = manifest.get("spec", {})
    page_key = str(spec.get("page_key", "")).strip().lower()
    points = spec.get("points", [])

    lead = ""
    if points:
        lead = str(points[0].get("text", "")).strip()
    if not lead:
        lead = "Strong women choose calm consistency, not chaos."
    if "||" in lead:
        hook, _ = lead.split("||", 1)
        lead = hook.strip()

    lead = re.sub(r"\s+", " ", lead)
    if len(lead) > 120:
        lead = lead[:117].rstrip() + "..."

    pool = list(dict.fromkeys(_load_hashtag_pool(project_root, page_key)))
    random.shuffle(pool)
    selected: list[str] = []
    seen: set[str] = set()
    for tag in pool:
        low = tag.lower()
        if low not in seen:
            selected.append(tag)
            seen.add(low)
        if len(selected) >= hashtag_count:
            break

    return f"{lead}\n\n{' '.join(selected)}"

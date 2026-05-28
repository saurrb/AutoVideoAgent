from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
import sqlite3

from openpyxl import load_workbook


@dataclass
class ContentBatch:
    heading_line1: str
    heading_line2: str
    cta: str
    rows: list[dict]

    @property
    def ids(self) -> list[int]:
        return [int(x["id"]) for x in self.rows]

    @property
    def batch_key(self) -> str:
        ids = self.ids
        return f"{ids[0]}_{ids[-1]}"


def _cell(v, default=""):
    return default if v is None else v


def _mark_used(conn: sqlite3.Connection, page_key: str, ids: list[int]) -> None:
    now = datetime.now(timezone.utc).isoformat()
    conn.executemany(
        "INSERT INTO content_items(page_key,item_id,used,used_at) VALUES(?,?,1,?) "
        "ON CONFLICT(page_key,item_id) DO UPDATE SET used=1, used_at=excluded.used_at",
        [(page_key, i, now) for i in ids],
    )
    conn.executemany(
        "INSERT INTO used_state(page_key,content_id,used,used_at) VALUES(?,?,1,?) "
        "ON CONFLICT(page_key,content_id) DO UPDATE SET used=1, used_at=excluded.used_at",
        [(page_key, i, now) for i in ids],
    )
    conn.commit()


def _already_used(conn: sqlite3.Connection, page_key: str, item_id: int) -> bool:
    row = conn.execute(
        "SELECT used FROM content_items WHERE page_key=? AND item_id=?",
        (page_key, item_id),
    ).fetchone()
    if bool(row and row[0] == 1):
        return True
    row2 = conn.execute(
        "SELECT used FROM used_state WHERE page_key=? AND content_id=?",
        (page_key, item_id),
    ).fetchone()
    return bool(row2 and row2[0] == 1)


def _sync_normalized_row(conn: sqlite3.Connection, page_key: str, row: dict) -> None:
    now = datetime.now(timezone.utc).isoformat()
    conn.execute(
        "INSERT INTO content_rows(page_key,content_id,heading,hook,reason,scene_prompt,updated_at) "
        "VALUES(?,?,?,?,?,?,?) "
        "ON CONFLICT(page_key,content_id) DO UPDATE SET "
        "heading=excluded.heading,hook=excluded.hook,reason=excluded.reason,scene_prompt=excluded.scene_prompt,updated_at=excluded.updated_at",
        (
            page_key,
            int(row.get("id", 0)),
            str(row.get("heading_line2", "")),
            str(row.get("hook", "")),
            str(row.get("reason", "")),
            str(row.get("scene_prompt", "")),
            now,
        ),
    )
    conn.execute(
        "INSERT INTO captions(page_key,content_id,caption,updated_at) VALUES(?,?,?,?) "
        "ON CONFLICT(page_key,content_id) DO UPDATE SET caption=excluded.caption, updated_at=excluded.updated_at",
        (page_key, int(row.get("id", 0)), str(row.get("caption", "")), now),
    )
    conn.execute(
        "INSERT INTO hashtags(page_key,content_id,hashtags,updated_at) VALUES(?,?,?,?) "
        "ON CONFLICT(page_key,content_id) DO UPDATE SET hashtags=excluded.hashtags, updated_at=excluded.updated_at",
        (page_key, int(row.get("id", 0)), str(row.get("hashtags", "")), now),
    )
    conn.execute(
        "INSERT INTO scene_prompts(page_key,content_id,scene_prompt,updated_at) VALUES(?,?,?,?) "
        "ON CONFLICT(page_key,content_id) DO UPDATE SET scene_prompt=excluded.scene_prompt, updated_at=excluded.updated_at",
        (page_key, int(row.get("id", 0)), str(row.get("scene_prompt", "")), now),
    )


def take_next_batch_from_excel(
    conn: sqlite3.Connection,
    page_key: str,
    xlsx_path: Path,
    sheet_name: str,
    batch_size: int = 5,
) -> ContentBatch:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    idx = {name: i + 1 for i, name in enumerate(header)}
    old_schema = [
        "id",
        "heading_line1",
        "heading_line2",
        "point_text",
        "highlight_first_words",
        "cta",
        "used",
    ]
    new_schema = ["Number", "Heading", "Hook", "Reason", "used"]
    has_old_schema = all(k in idx for k in old_schema)
    has_new_schema = all(k in idx for k in new_schema)
    if not has_old_schema and not has_new_schema:
        raise ValueError(
            f"Unsupported columns in {xlsx_path}. "
            f"Need either {old_schema} or {new_schema}."
        )

    candidates: list[dict] = []
    for r in range(2, ws.max_row + 1):
        if has_new_schema:
            item_id = int(_cell(ws.cell(r, idx["Number"]).value, 0))
        else:
            item_id = int(_cell(ws.cell(r, idx["id"]).value, 0))
        if item_id <= 0:
            continue
        excel_used = int(_cell(ws.cell(r, idx["used"]).value, 0)) == 1
        db_used = _already_used(conn, page_key, item_id)
        if excel_used or db_used:
            continue
        if has_new_schema:
            heading = str(_cell(ws.cell(r, idx["Heading"]).value)).strip()
            hook = str(_cell(ws.cell(r, idx["Hook"]).value)).strip()
            reason = str(_cell(ws.cell(r, idx["Reason"]).value)).strip()
            point_text = f"{hook}||{reason}" if reason else hook
            candidates.append(
                {
                    "row_num": r,
                    "id": item_id,
                    "heading_line1": "DAILY DESIRE FACTS",
                    "heading_line2": heading or "DESIRE FACT",
                    "point_text": point_text,
                    "highlight_first_words": 3,
                    "cta": "",
                    "scene_prompt": str(_cell(ws.cell(r, idx.get("scene_prompt", 0)).value)).strip()
                    if idx.get("scene_prompt")
                    else "",
                    "hook": hook,
                    "reason": reason,
                    "caption": str(_cell(ws.cell(r, idx.get("caption", 0)).value)).strip() if idx.get("caption") else "",
                    "hashtags": str(_cell(ws.cell(r, idx.get("hashtags", 0)).value)).strip() if idx.get("hashtags") else "",
                }
            )
        else:
            candidates.append(
                {
                    "row_num": r,
                    "id": item_id,
                    "heading_line1": str(_cell(ws.cell(r, idx["heading_line1"]).value)).strip(),
                    "heading_line2": str(_cell(ws.cell(r, idx["heading_line2"]).value)).strip(),
                    "point_text": str(_cell(ws.cell(r, idx["point_text"]).value)).strip(),
                    "highlight_first_words": int(_cell(ws.cell(r, idx["highlight_first_words"]).value, 3)),
                    "cta": str(_cell(ws.cell(r, idx["cta"]).value)).strip(),
                }
            )
        if len(candidates) == batch_size:
            break

    if len(candidates) < batch_size:
        raise ValueError(f"Not enough unused items for page '{page_key}'. Need {batch_size}.")

    heading_line1 = candidates[0]["heading_line1"]
    heading_line2 = candidates[0]["heading_line2"]
    cta = candidates[0]["cta"]
    ids = [x["id"] for x in candidates]

    for row in candidates:
        ws.cell(row["row_num"], idx["used"], 1)
    wb.save(xlsx_path)
    _mark_used(conn, page_key, ids)
    for row in candidates:
        _sync_normalized_row(conn, page_key, row)
    conn.commit()

    return ContentBatch(
        heading_line1=heading_line1,
        heading_line2=heading_line2,
        cta=cta,
        rows=candidates,
    )


def import_excel_to_db(
    conn: sqlite3.Connection,
    page_key: str,
    xlsx_path: Path,
    sheet_name: str,
) -> int:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    idx = {name: i + 1 for i, name in enumerate(header)}
    required = ["id", "heading_line1", "heading_line2", "point_text", "highlight_first_words", "cta"]
    if not all(k in idx for k in required):
        raise ValueError(f"Expected old-schema columns {required} in {xlsx_path}")
    now = datetime.now(timezone.utc).isoformat()
    upserts: list[tuple] = []
    for r in range(2, ws.max_row + 1):
        item_id = int(_cell(ws.cell(r, idx["id"]).value, 0))
        if item_id <= 0:
            continue
        heading_line1 = str(_cell(ws.cell(r, idx["heading_line1"]).value)).strip()
        heading_line2 = str(_cell(ws.cell(r, idx["heading_line2"]).value)).strip()
        point_text = str(_cell(ws.cell(r, idx["point_text"]).value)).strip()
        hfw = int(_cell(ws.cell(r, idx["highlight_first_words"]).value, 3))
        cta = str(_cell(ws.cell(r, idx["cta"]).value)).strip()
        scene_prompt = str(_cell(ws.cell(r, idx.get("scene_prompt", 0)).value)).strip() if idx.get("scene_prompt") else ""
        caption = str(_cell(ws.cell(r, idx.get("caption", 0)).value)).strip() if idx.get("caption") else ""
        hashtags = str(_cell(ws.cell(r, idx.get("hashtags", 0)).value)).strip() if idx.get("hashtags") else ""
        upserts.append(
            (
                page_key,
                item_id,
                heading_line1,
                heading_line2,
                point_text,
                hfw,
                cta,
                scene_prompt,
                caption,
                hashtags,
                now,
            )
        )
    conn.executemany(
        "INSERT INTO content_bank_rows("
        "page_key,item_id,heading_line1,heading_line2,point_text,highlight_first_words,cta,scene_prompt,caption,hashtags,updated_at"
        ") VALUES(?,?,?,?,?,?,?,?,?,?,?) "
        "ON CONFLICT(page_key,item_id) DO UPDATE SET "
        "heading_line1=excluded.heading_line1,heading_line2=excluded.heading_line2,point_text=excluded.point_text,"
        "highlight_first_words=excluded.highlight_first_words,cta=excluded.cta,scene_prompt=excluded.scene_prompt,"
        "caption=excluded.caption,hashtags=excluded.hashtags,updated_at=excluded.updated_at",
        upserts,
    )
    conn.commit()
    return len(upserts)


def take_next_batch_from_db(
    conn: sqlite3.Connection,
    page_key: str,
    batch_size: int = 5,
) -> ContentBatch:
    rows = conn.execute(
        "SELECT item_id, heading_line1, heading_line2, point_text, highlight_first_words, cta, scene_prompt, caption, hashtags "
        "FROM content_bank_rows "
        "WHERE page_key=? "
        "AND item_id NOT IN (SELECT item_id FROM content_items WHERE page_key=? AND used=1) "
        "AND item_id NOT IN (SELECT content_id FROM used_state WHERE page_key=? AND used=1) "
        "ORDER BY item_id ASC "
        "LIMIT ?",
        (page_key, page_key, page_key, batch_size),
    ).fetchall()
    if len(rows) < batch_size:
        raise ValueError(f"Not enough unused DB items for page '{page_key}'. Need {batch_size}.")

    candidates: list[dict] = []
    for r in rows:
        candidates.append(
            {
                "id": int(r[0]),
                "heading_line1": str(r[1] or "").strip(),
                "heading_line2": str(r[2] or "").strip(),
                "point_text": str(r[3] or "").strip(),
                "highlight_first_words": int(r[4] or 3),
                "cta": str(r[5] or "").strip(),
                "scene_prompt": str(r[6] or "").strip(),
                "caption": str(r[7] or "").strip(),
                "hashtags": str(r[8] or "").strip(),
            }
        )

    ids = [x["id"] for x in candidates]
    _mark_used(conn, page_key, ids)
    conn.commit()

    return ContentBatch(
        heading_line1=candidates[0]["heading_line1"],
        heading_line2=candidates[0]["heading_line2"],
        cta=candidates[0]["cta"],
        rows=candidates,
    )
